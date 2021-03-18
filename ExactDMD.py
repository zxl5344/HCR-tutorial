from openpyxl import load_workbook
import numpy as np
from numpy import linalg as LA
import pandas as pd

#read data from excel
wb = load_workbook(filename = '#P_car_data.xlsx')
name = wb.sheetnames[0]
sheet_ranges = wb[name]						#first sheets of excel
df = pd.DataFrame(sheet_ranges.values)
arr = df.to_numpy()	

def Exact_DMD(X,Y):
	#find SVD
	U,s,Vh = LA.svd(X, full_matrices=False)      		#Vh is the complex conjugate; full_matrices is disabled

	V = np.conjugate(Vh)
	V = V.T					 
	Uh = np.conjugate(U)					
	Uh = Uh.T						#Uh is the complex conjugate of U

	s = np.diag(s)			 			#resize the s		
	s_inv = LA.inv(s)					#s_inv is the inverse of s

	#define Nex matrix
	A_Tilde = Uh@Y@V@s_inv

	#eig of A_Tilde 
	lamda,omega = LA.eig(A_Tilde)

	#define phi
	for i in range(0,len(lamda)): 
		omega[:,i] *=lamda[i]

	phi = Y@V@LA.inv(np.diag(lamda))@omega
	return phi

#arrange the data
for i in range(0,30):
	X = arr[(1+i*1802):(1802+i*1802),2:6]			#Take 1(z_0) to 1801(z_m-1) rows and 2 to 5 columns as X matrix
	Y = arr[(2+i*1802):(1803+i*1802),2:6]			#Take 2(z_0) to 1802(z_m) rows and 2 to 5 columns as Y matrix
	X = X.astype(np.float64)  				#change type of X from object to float64
	Y = Y.astype(np.float64)
	X = X.T							#X.T means tranpose of X
	Y = Y.T
	if i == 0:
		X1 = X
		Y1 = Y
	else:
		X1 = np.append(X1,X,axis = 1)
		Y1 = np.append(Y1,Y,axis = 1)

phi = Exact_DMD(X1,Y1)
print(phi)
